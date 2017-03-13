import openpyxl
from openpyxl import load_workbook
import re
import numpy as np


example1 = ['MACS30200', \
            'Perspectives on Advanced Computational Topics', \
            'Saieh Hall for Economics 247', \
            'Mon Wed : 11:30 AM-12:50 PM']
example2 = ['MATH20300', \
            'Analysis In Rn-1',\
            'Eckhart Hall 202',\
            'Mon Wed Fri : 12:30 PM-01:20 PM']
example3 = ['MPCS58020', \
            'Time Series Analysis and Stochastic Processes', \
            'TBA', \
            'Mon : 05:30 PM-08:30 PM']
example4 = ['LAWS90213', \
            'Mental Health Advocacy Clinic', \
            'Laird Bell Quadrangle B', \
            'Thu : 04:00 PM-06:00 PM']

wb =load_workbook(filename = 'template.xlsx')
template = wb.active

# for i in range(3, 63):
#    index = 'A' + str(i)
#    print(template[index].value)

def time_calculator(time):
    '''
    function converts time, for example: 13:30 to 13.5
    '''

    hours = int(time)
    hours_decimal = round(time - int(time), 1)
    mins = hours_decimal * 100 / 60
    hours_float = hours + mins
    return hours_float
    

# get the number of cells spans of that particular course
def course_info(time):
    '''
    input: time = '12:30 PM-01:00 PM'
    return: (12.5, 13.0, 1.0, 1)
    '''
    span_nums_str = re.findall('\d+', time)
    end_time = float(span_nums_str[-2] + '.' + span_nums_str[-1])
    start_time = float(span_nums_str[0] + '.' + span_nums_str[1])
    AM_PM = re.findall('(AM|PM)', time)
    days = re.findall('(Mon|Tue|Wed|Thu|Fri|Sat|Sun)', time)

    if AM_PM[0] == 'PM' and int(start_time) != 12:
        start_time += 12
    if AM_PM[-1] == 'PM' and int(end_time) != 12:
        end_time += 12

    # time conversion
    end_time = time_calculator(end_time)
    start_time = time_calculator(start_time)

    # convert to hours
    num_hours = end_time - start_time
    # convert to minutes
    num_mins = round(num_hours * 60, 0)
    # convert to 30mins
    num_15mins = num_mins / 15

    #special case
    if num_15mins - int(num_15mins) != 0:
        num_15mins += 1
    num_cells = round(num_15mins, 0)

    return days, start_time, end_time, num_mins, num_cells



# get a list of all time slots
# ['A3',
#  'A5',
#  'A7',
#  'A9',
#  'A11',
#  'A13',
#  'A15',
#  'A17',
#  'A19',
#  'A21',
#  'A23',
#  'A25',
#  'A27',
#  'A29',
#  'A31',
#  'A33',
#  'A35',
#  'A37',
#  'A39',
#  'A41',
#  'A43',
#  'A45',
#  'A47',
#  'A49',
#  'A51',
#  'A53',
#  'A55',
#  'A57',
#  'A59']
time_cells = []
for index in range(3, 60, 2):
    time_cells.append('A' + str(index))

# get a list of all possible time slots
# ['07:00 AM-07:30 AM ',
#  '07:30 AM-08:00 AM',
#  '08:00 AM-08:30 AM ',
#  '08:30 AM-09:00 AM',
#  '09:00 AM-09:30 AM',
#  '09:30 AM-10:00 AM',
#  '10:00 AM-10:30 AM ',
#  '10:30 AM-11:00 AM',
#  '11:00 AM-11:30 AM',
#  '11:30 AM-12:00 PM',
#  '12:00 PM-12:30 PM ',
#  '12:30 PM-01:00 PM',
#  '01:00 PM-01:30 PM ',
#  '01:30 PM-02:00 PM',
#  '02:00 PM-02:30 PM ',
#  '02:30 PM-03:00 PM',
#  '03:00 PM-03:30 PM ',
#  '03:30 PM-04:00 PM',
#  '04:00 PM-04:30 PM ',
#  '04:30 PM-05:00 PM',
#  '05:00 PM-05:30 PM ',
#  '05:30 PM-06:00 PM',
#  '06:00 PM-06:30 PM ',
#  '06:30 PM-07:00 PM',
#  '07:00 PM-07:30 PM ',
#  '07:30 PM-08:00 PM',
#  '08:00 PM-08:30 PM ',
#  '08:30 PM-09:00 PM',
#  '09:00 PM-10:00 PM ']
time_slots = []
for cells in time_cells:
    time_slots.append(template[cells].value)

################### Matching days to columns ###################
days_map = {'Mon':'B', 'Tue':'C', 'Wed':'D', \
            'Thu':'E', 'Fri':'F', 'Sat':'G', 'Sun':'H'}

################## Matching times to rows ###################
slots_map = {}
for slots, index in zip(time_slots, range(3, 60, 2)): 
    days, start_time, end_time, \
            num_mins, num_cells = course_info(slots)
    slots_map[start_time] = index # need to count title rows


def merger_pattern(days, start_time, num_cells):
    cols = []
    for day in days:
        cols.append(days_map[day])
    
    starting_row= int(slots_map[start_time])
    ending_row = int(starting_row + num_cells - 1)

    merge_patterns = []
    starting_cells = []
    for col in cols:
        starting_cells.append(col+str(starting_row))
        merge_patterns.append(col+str(starting_row)\
                                +':'+col+str(ending_row))

    return starting_cells, merge_patterns


def write_schedule(template, starting_cells,\
                    merge_patterns, course):
    ccn, name, loc, time = course
    days, start_time, end_time, num_mins, \
            num_cells = course_info(time)
    content = name + '\n' + '\n' + 'Location: ' + loc
    # merge cells and write it
    for start, pattern in zip(starting_cells, merge_patterns):
        template.merge_cells(pattern)
        template[start] = content



def builder(course_list, file_name):
    wb =load_workbook(filename = 'template.xlsx')
    template = wb.active

    for course in course_list:
        ccn, name, loc, time = course
        days, start_time, end_time, \
                num_mins, num_cells = course_info(time)
        starting_cells, merge_patterns = merger_pattern(days, \
                                                        start_time,\
                                                        num_cells)
        write_schedule(template, starting_cells, \
                       merge_patterns, course)

    wb.save(file_name)
